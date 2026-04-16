# -*- coding: utf-8 -*-
"""
Excel 导出功能
========================
负责将用例数据导出为 Excel 文件

职责：
- 将 cases 规范化为 Excel 行
- 写入 file_store
- 返回下载所需元数据

支持：
1) 新旧字段兼容：
   - 新字段（优先）：用例名称 / 所属模块 / 前置条件 / 步骤描述 / 预期结果 / 标签 / 用例等级 / 用例状态 / 编辑模式 / 备注
   - 旧字段（兼容）：title / module / precondition / steps / expected / tags / priority / status / remark
2) 模块规则：
   - Excel“所属模块”固定写 requirement_id
3) 标签规则：
   - 用例等级=P0 -> 标签固定=冒烟测试
   - 其他 -> 标签固定=功能测试
4) steps/expected 结构化：
   - 自动编号（1. 2. 3.）
   - 尽量保持 steps 与 expected 一一对应
   - 同时兼容字符串和 list
5) 责任人支持：
   - raw["owner"] -> 写入 Excel “责任人”列
6) 编辑模式支持：
   - 固定输出 STEP（若明确传入“更新”则保留）
7) 不影响旧导出链路
"""

from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
import uuid
import io
import re
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from app.services import file_store


HEADERS = [
    "用例名称", "所属模块", "标签", "前置条件", "步骤描述",
    "预期结果", "编辑模式", "备注", "用例状态", "责任人",
    "用例等级", "是否可自动化", "是否已自动化"
]

EXCEL_CELL_LIMIT = 32000

DEFAULT_TAG_SMOKE = "冒烟测试"
DEFAULT_TAG_FUNCTION = "功能测试"

WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
TOP_ALIGNMENT = Alignment(vertical="top")


def _normalize_text(v: Any) -> str:
    return str(v or "").strip()


def _truncate(text: str) -> str:
    if not text:
        return ""
    if len(text) > EXCEL_CELL_LIMIT:
        return text[:EXCEL_CELL_LIMIT] + "\n【内容过长，已截断】"
    return text


def _cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, list):
        return _truncate("\n".join(str(x) for x in v if str(x).strip()))
    return _truncate(str(v))


def _pick_first(raw: Dict[str, Any], *keys: str, default: Any = "") -> Any:
    for k in keys:
        if k in raw and raw.get(k) is not None and str(raw.get(k)).strip() != "":
            return raw.get(k)
    return default


def _normalize_priority(raw: Any) -> str:
    """
    归一优先级到 P0/P1/P2
    """
    if raw is None:
        return "P1"

    s = str(raw).strip()
    if not s:
        return "P1"

    s_up = s.upper()

    if s_up in {"0", "1", "2"}:
        return f"P{s_up}"

    if s_up in {"P0", "P1", "P2"}:
        return s_up

    m = re.search(r"P\s*([0-9]+)", s_up)
    if m:
        n = m.group(1)
        if n in {"0", "1", "2"}:
            return f"P{n}"

    m2 = re.search(r"([0-9]+)", s_up)
    if m2 and m2.group(1) in {"0", "1", "2"}:
        return f"P{m2.group(1)}"

    if "高" in s or "critical" in s_up or "high" in s_up:
        return "P0"
    if "低" in s or "low" in s_up:
        return "P2"

    return "P1"


def _tag_from_priority(priority: Any) -> str:
    return DEFAULT_TAG_SMOKE if _normalize_priority(priority) == "P0" else DEFAULT_TAG_FUNCTION


def _normalize_automatable(raw: Any) -> str:
    s = str(raw or "").strip().lower()
    if s in {"是", "true", "1", "y", "yes"}:
        return "是"
    if s in {"否", "false", "0", "n", "no"}:
        return "否"
    return "否"


def _normalize_edit_mode(raw: Any) -> str:
    """
    编辑模式默认固定 STEP
    """
    s = str(raw or "").strip()
    if s == "更新":
        return "更新"
    return "STEP"


def _normalize_case_status(raw: Any) -> str:
    s = str(raw or "").strip()
    if not s:
        return "未开始"

    s_lower = s.lower()
    mapping = {
        "未开始": "未开始",
        "待执行": "未开始",
        "待开始": "未开始",
        "pending": "未开始",
        "running": "执行中",
        "done": "已执行",
        "finished": "已执行",
        "deprecated": "已废弃",
    }
    return mapping.get(s_lower, s)


def _strip_leading_index(s: str) -> str:
    """
    去掉开头可能存在的编号：
    - 1. xxx
    - 1、xxx
    - 1) xxx
    - (1) xxx
    """
    if not s:
        return ""
    s = str(s).strip()
    s = re.sub(r"^\(?\s*\d+\s*\)?\s*[\.\、\)\-:：]\s*", "", s)
    return s.strip()


def _clean_line_text(s: str) -> str:
    s = _strip_leading_index(str(s or "").strip())
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _split_multiline_to_list(v: Any) -> List[str]:
    """
    兼容：
    - list
    - 多行字符串
    - 单行字符串
    """
    if v is None:
        return []

    if isinstance(v, list):
        result: List[str] = []
        for item in v:
            item_str = str(item).strip()
            if not item_str:
                continue
            for line in item_str.splitlines():
                line = _clean_line_text(line)
                if line:
                    result.append(line)
        return result

    if isinstance(v, str):
        text = v.strip()
        if not text:
            return []

        lines = []
        for line in text.splitlines():
            line = _clean_line_text(line)
            if line:
                lines.append(line)

        if lines:
            return lines

        text = _clean_line_text(text)
        return [text] if text else []

    s = _clean_line_text(str(v).strip())
    return [s] if s else []


def _dedup_keep_order(items: List[str]) -> List[str]:
    seen = set()
    result: List[str] = []
    for item in items:
        key = item.strip()
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(item)
    return result


def _format_indexed(lines: List[str], style: str = "dot") -> List[str]:
    out: List[str] = []
    clean_lines = [_clean_line_text(x) for x in lines if _clean_line_text(x)]

    for i, x2 in enumerate(clean_lines):
        if style == "cn":
            out.append(f"{i + 1}、{x2}")
        else:
            out.append(f"{i + 1}. {x2}")
    return out


def _is_generic_expected(text: str) -> bool:
    s = _clean_line_text(text)
    if not s:
        return True

    bad_patterns = (
        "符合业务规则",
        "符合预期",
        "页面正常",
        "结果正确",
        "功能正常",
        "状态正常",
        "显示正常",
        "无异常即可",
        "成功即可",
        "系统处理结果符合预期",
        "步骤执行完成后",
        "结果应符合业务规则",
        "正常即可",
        "无问题",
    )
    return any(x in s for x in bad_patterns)


def _build_expected_from_steps(steps: List[str]) -> List[str]:
    """
    当 expected 缺失或明显不足时，根据 steps 自动兜底生成更专业的预期。
    """
    result: List[str] = []

    for step in steps:
        s = _clean_line_text(step)

        if s.startswith("进入"):
            result.append("目标页面进入成功，页面展示完整且无异常报错")
        elif s.startswith("打开"):
            result.append("对应面板、弹窗或菜单打开成功，内容展示完整")
        elif s.startswith("点击"):
            result.append("点击操作被正确响应，页面状态或业务结果按预期变化")
        elif s.startswith("选择"):
            result.append("选项选择成功，所选内容正确生效并展示当前选中状态")
        elif s.startswith("输入"):
            result.append("输入内容被正确接收并展示，字段格式与内容保持正确")
        elif s.startswith("提交"):
            result.append("提交操作执行完成，系统返回结果正确并给出明确反馈")
        elif s.startswith("切换"):
            result.append("切换操作执行成功，目标状态、样式或数据正确更新")
        elif s.startswith("查看") or s.startswith("观察"):
            result.append("页面展示、数据内容和状态变化与预期一致")
        elif s.startswith("校验") or s.startswith("确认"):
            result.append("校验结果符合业务规则，相关字段和状态保持正确")
        elif s.startswith("刷新"):
            result.append("刷新后页面和数据加载正常，状态保持正确")
        elif s.startswith("保存"):
            result.append("保存成功，保存结果可见且数据持久化正确")
        elif s.startswith("删除"):
            result.append("删除成功，目标数据被正确移除且页面状态同步更新")
        elif s.startswith("新增"):
            result.append("新增成功，新数据正确写入并在页面中可见")
        elif s.startswith("查询"):
            result.append("查询结果返回正确，结果集与查询条件保持一致")
        elif s.startswith("上传"):
            result.append("上传成功，文件或数据被正确接收并展示处理结果")
        elif s.startswith("下载"):
            result.append("下载成功，文件内容、格式和命名符合预期")
        elif s.startswith("悬浮"):
            result.append("悬浮交互被正确响应，提示信息展示完整且内容正确")
        elif s.startswith("拖拽"):
            result.append("拖拽操作响应正常，界面或数据随操作正确变化")
        elif s.startswith("放大") or s.startswith("缩小"):
            result.append("缩放操作响应正常，图表或页面内容展示连续且无异常")
        elif s.startswith("返回"):
            result.append("返回操作执行成功，页面跳转和状态保持正确")
        else:
            result.append("操作执行完成后，系统应返回明确结果且页面状态正确更新")

    return _dedup_keep_order(result)


def _clean_remark_text(text: Any) -> str:
    s = str(text or "").strip()
    if not s:
        return ""

    bad_patterns = (
        "根据审核意见",
        "根据测试点生成",
        "AI兜底生成",
        "已按专业测试用例格式",
        "重组步骤与预期",
        "根据审核结果优化",
        "根据审核意见修正",
        "根据测试点生成：",
        "已按专业测试用例",
    )

    if any(x in s for x in bad_patterns):
        return ""

    return s


def _normalize_steps_expected(
    steps_raw: Any,
    expected_raw: Any,
    *,
    index_style: str = "dot",
) -> Tuple[List[str], List[str]]:
    """
    尽量保持 steps / expected 一一对应：
    - 都归一为 list[str]
    - 自动编号
    - expected 不足补齐
    - expected 过多则截断
    - 过滤笼统预期
    """
    steps = _dedup_keep_order(_split_multiline_to_list(steps_raw))
    expected = _dedup_keep_order(_split_multiline_to_list(expected_raw))

    expected = [x for x in expected if not _is_generic_expected(x)]

    if not steps and expected:
        steps = expected
        expected = []

    if not steps and not expected:
        return [], []

    if not steps:
        steps = ["执行对应业务操作"]

    if len(expected) < len(steps):
        generated = _build_expected_from_steps(steps)
        for i in range(len(expected), len(steps)):
            if i < len(generated):
                expected.append(generated[i])
            else:
                expected.append("操作执行完成后，系统应返回明确结果且页面状态正确更新")

    if len(expected) > len(steps):
        expected = expected[:len(steps)]

    steps_idx = _format_indexed(steps, style=index_style)
    expected_idx = _format_indexed(expected, style=index_style)

    return steps_idx, expected_idx


def normalize_case(
    raw: Dict[str, Any],
    default_module: str,
    *,
    requirement_id: Optional[str] = None,
) -> Optional[Dict[str, Any]]:
    """
    新旧字段兼容：
    新字段优先，旧字段兜底

    规则：
    - 所属模块固定 = requirement_id（若传入）
    - 标签固定由优先级推导：P0=冒烟测试，其余=功能测试
    """
    if not isinstance(raw, dict):
        return None

    case_name = _normalize_text(
        _pick_first(raw, "用例名称", "title", "name", default="")
    )
    if not case_name:
        return None

    rid = _normalize_text(requirement_id or default_module) or "整体功能"
    module = rid

    priority = _normalize_priority(
        _pick_first(raw, "用例等级", "priority", default="P1")
    )

    tags = _tag_from_priority(priority)

    precondition = _normalize_text(
        _pick_first(raw, "前置条件", "precondition", "preconditions", default="")
    )

    steps_raw = _pick_first(raw, "步骤描述", "steps", default=[])
    expected_raw = _pick_first(raw, "预期结果", "expected", default=[])

    steps_idx, expected_idx = _normalize_steps_expected(
        steps_raw,
        expected_raw,
        index_style="dot",
    )

    owner = _normalize_text(raw.get("owner", ""))
    remark = _clean_remark_text(
        _pick_first(raw, "备注", "remark", "comments", default="")
    )
    automatable = _normalize_automatable(
        _pick_first(raw, "是否可自动化", "automatable", default="否")
    )
    edit_mode = _normalize_edit_mode(
        _pick_first(raw, "编辑模式", "edit_mode", default="STEP")
    )
    case_status = _normalize_case_status(
        _pick_first(raw, "用例状态", "status", default="待执行")
    )

    case = {
        "case_name": case_name,
        "module": module,
        "tags": tags,
        "precondition": precondition,
        "steps": steps_idx,
        "expected": expected_idx,
        "edit_mode": edit_mode,
        "remark": remark,
        "status": case_status,
        "priority": priority,
        "automatable": automatable,
        "owner": owner,
    }

    return case


def _store_excel_bytes(content: bytes) -> str:
    """
    把 bytes 写入 file_store，返回 file_id（必返回，失败抛异常）
    """
    save_bytes = getattr(file_store, "save_bytes", None)
    if callable(save_bytes):
        return str(save_bytes(content, file_id=None, ext="xlsx"))

    fid = uuid.uuid4().hex
    save_file = getattr(file_store, "save_file", None)
    if not callable(save_file):
        raise RuntimeError("file_store has no save_bytes/save_file")

    try:
        save_file(fid, content, ext="xlsx")  # type: ignore
    except TypeError:
        save_file(fid, content)  # type: ignore

    return fid


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_body_row(ws, row: int) -> None:
    for col_idx in range(1, len(HEADERS) + 1):
        ws.cell(row=row, column=col_idx).alignment = TOP_ALIGNMENT

    for col in ("E", "F", "G", "H", "I", "J", "K"):
        ws[f"{col}{row}"].alignment = WRAP_ALIGNMENT


def _set_column_widths(ws) -> None:
    widths = {
        "A": 14,   # ID
        "B": 40,   # 用例名称
        "C": 26,   # 所属模块
        "D": 14,   # 标签
        "E": 34,   # 前置条件
        "F": 68,   # 步骤描述
        "G": 68,   # 预期结果
        "H": 12,   # 编辑模式
        "I": 34,   # 备注
        "J": 12,   # 用例状态
        "K": 16,   # 责任人
        "L": 10,   # 用例等级
        "M": 12,   # 是否可自动化
        "N": 12,   # 是否已自动化
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _set_row_height_if_needed(ws, row: int) -> None:
    texts = []
    for col in ("E", "F", "G", "I"):
        texts.append(str(ws[f"{col}{row}"].value or ""))

    max_lines = 1
    for text in texts:
        max_lines = max(max_lines, text.count("\n") + 1)

    ws.row_dimensions[row].height = min(max(22, max_lines * 18), 180)


def export_cases_to_excel(
    cases: List[Dict[str, Any]],
    requirement_id: str,
    *,
    filename: Optional[str] = None,
) -> Dict[str, Any]:
    """
    导出用例数据到 Excel，并写入 file_store。

    返回字段：
    - download_ready: bool
    - file_id: str | None
    - file_name: str
    - download_url: str | None
    - total_cases: int
    - error: str | None
    """
    rid = (requirement_id or "").strip() or "UNKNOWN_REQUIREMENT"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    file_name = (filename or f"测试用例_{rid}_{timestamp}.xlsx").strip()
    if not file_name.lower().endswith(".xlsx"):
        file_name += ".xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    ws.append(HEADERS)
    ws.freeze_panes = "A2"

    _style_header(ws)

    written = 0
    owner_for_trace = ""

    for raw in cases or []:
        case = normalize_case(raw, default_module=rid, requirement_id=rid)
        if not case:
            continue

        if not owner_for_trace:
            owner_for_trace = _normalize_text(case.get("owner", ""))

        ws.append([
            _cell(case["case_name"]),
            _cell(rid),                      # ⭐ 所属模块固定 requirement_id
            _cell(_tag_from_priority(case["priority"])),  # ⭐ P0=冒烟测试，其余=功能测试
            _cell(case["precondition"]),
            _cell(case["steps"]),
            _cell(case["expected"]),
            _cell(case["edit_mode"]),
            _cell(case["remark"]),
            _cell(case["status"]),
            _cell(case.get("owner", "")),
            _cell(case["priority"]),
            _cell(case["automatable"]),
            "否",
        ])

        row = ws.max_row
        _style_body_row(ws, row)
        _set_row_height_if_needed(ws, row)
        written += 1

    _set_column_widths(ws)
    ws.row_dimensions[1].height = 24

    file_id: Optional[str] = None
    export_error: Optional[str] = None

    try:
        bio = io.BytesIO()
        wb.save(bio)
        content = bio.getvalue()
        file_id = _store_excel_bytes(content)
    except Exception as e:
        file_id = None
        export_error = repr(e)

    download_ready = bool(file_id)

    if download_ready:
        if owner_for_trace:
            download_url = (
                f"/testcase/download?file_id={file_id}"
                f"&filename={quote(file_name)}"
                f"&owner={quote(owner_for_trace)}"
            )
        else:
            download_url = f"/testcase/download?file_id={file_id}&filename={quote(file_name)}"
    else:
        download_url = None

    return {
        "download_ready": download_ready,
        "file_id": file_id,
        "file_name": file_name,
        "download_url": download_url,
        "total_cases": written,
        "error": export_error,
    }