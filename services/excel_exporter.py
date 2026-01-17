#! /usr/bin/python3
# coding=utf-8

from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment

# ===============================
# Excel 表头
# ===============================
HEADERS = [
    "用例名称", "所属模块", "标签", "前置条件", "步骤描述",
    "预期结果", "编辑模式", "备注", "用例状态", "责任人",
    "用例等级", "是否可自动化", "是否已自动化"
]

EXCEL_CELL_LIMIT = 32000


# ===============================
# 工具函数
# ===============================
def _truncate(text: str) -> str:
    if len(text) > EXCEL_CELL_LIMIT:
        return text[:EXCEL_CELL_LIMIT] + "\n【内容过长，已截断】"
    return text


def _cell(v):
    if v is None:
        return ""
    if isinstance(v, list):
        return _truncate("\n".join(str(x) for x in v))
    return _truncate(str(v))


# ===============================
# 🔥 新增：steps 终极归一化（本次核心修复）
# ===============================
def _normalize_steps(steps) -> list[str]:
    """
    将 steps 统一为 List[str]
    支持：
    - str
    - List[str]
    - List[{"step": "..."}]
    - List[{"desc": "..."}]
    - 混合结构
    """
    if not steps:
        return []

    if isinstance(steps, str):
        return [s.strip() for s in steps.split("\n") if s.strip()]

    normalized = []

    for s in steps:
        if isinstance(s, str):
            normalized.append(s.strip())
        elif isinstance(s, dict):
            normalized.append(
                str(
                    s.get("step")
                    or s.get("desc")
                    or s.get("content")
                    or ""
                ).strip()
            )
        else:
            normalized.append(str(s).strip())

    return [x for x in normalized if x]


def clean_case_name(name: str) -> str:
    """
    Excel 专用：终极兜底清洗用例名称前缀
    """
    if not name:
        return name

    name = name.strip()

    # ① 冒号型前缀
    name = re.sub(r"^[A-Za-z0-9_-]+:\s*", "", name)

    # ② 找到第一个中文字符
    m = re.search(r"[\u4e00-\u9fff]", name)
    if m:
        name = name[m.start():]

    return name.strip()


def clean_module_name(module: str) -> str:
    """
    去掉模块名中的英文括号说明
    """
    if not module:
        return module

    return module.split(" (")[0].strip()


# ===============================
# 🔥 展开 case
# ===============================
def flatten_cases(raw_cases: list) -> list:
    result = []

    for raw in raw_cases:
        if isinstance(raw, dict) and isinstance(raw.get("test_cases"), list):
            for c in raw["test_cases"]:
                merged = dict(c)
                merged["module"] = raw.get("module", "")
                merged["test_point_name"] = raw.get("test_point_name", "")
                result.append(merged)
        else:
            result.append(raw)

    return result


# ===============================
# ⭐ Case 标准化
# ===============================
def normalize_case(raw: dict) -> dict:
    # ✅ 核心修复：steps 彻底归一
    steps = _normalize_steps(raw.get("steps"))

    expected = (
        raw.get("expected")
        or raw.get("expected_result")
        or raw.get("expected_results")
        or ""
    )
    if isinstance(expected, list):
        expected = "\n".join(str(x) for x in expected)

    precondition = (
        raw.get("precondition")
        or raw.get("preconditions")
        or ""
    )
    if isinstance(precondition, list):
        precondition = "\n".join(str(x) for x in precondition)

    case_name = (
        raw.get("title")
        or raw.get("case_name")
        or raw.get("name")
        or f"【{raw.get('type','')}】{raw.get('test_point_name','未命名用例')}"
    )

    case_name = clean_case_name(case_name)
    module = clean_module_name(raw.get("module", ""))

    return {
        "case_name": case_name,
        "module": module,
        "tags": "功能测试",
        "precondition": precondition,
        "steps": steps,
        "expected": expected,
        "priority": infer_priority(steps),
        "automatable": infer_automatable(steps),
    }


def infer_priority(steps):
    text = " ".join(steps)
    if any(k in text for k in ["资金", "下单", "风控"]):
        return "P0"
    if "异常" in text:
        return "P1"
    return "P2"


def infer_automatable(steps):
    return "是" if any("接口" in s for s in steps) else "否"


# ===============================
# ⭐ Excel 导出
# ===============================
def export_excel(raw_cases: list):
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    ws.append(HEADERS)
    ws.freeze_panes = "A2"

    wrap = Alignment(wrap_text=True, vertical="top")

    flat_cases = flatten_cases(raw_cases)

    for raw in flat_cases:
        c = normalize_case(raw)

        ws.append([
            _cell(c["case_name"]),
            _cell(c["module"]),
            c["tags"],
            _cell(c["precondition"]),
            _cell(c["steps"]),
            _cell(c["expected"]),
            "STEP",
            "",
            "未开始",
            "",
            c["priority"],
            c["automatable"],
            "否"
        ])

        row = ws.max_row
        ws[f"E{row}"].alignment = wrap
        ws[f"F{row}"].alignment = wrap

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 70
    ws.column_dimensions["F"].width = 60

    filename = f"测试用例_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename
